import openpyxl

def new_admision(name,fname,mname):
    wb=openpyxl.load_workbook("data.xlsx")
    st1=wb["Sheet1"]
    rows=st1.max_row+1
    id =int(st1.cell(rows-1,1).value)
    st1.cell(rows,1).value=id+1
    st1.cell(rows,2).value=name
    st1.cell(rows,3).value=fname
    st1.cell(rows,4).value=mname
    wb.save("data.xlsx")

def deleterow(row):
    wb=openpyxl.load_workbook("data.xlsx")
    st1=wb["Sheet1"]
    st1.cell(row,1).value=None
    st1.cell(row,2).value=None
    st1.cell(row,3).value=None
    st1.cell(row,4).value=None
    wb.save("data.xlsx")

def removebyid(idval):
    wb=openpyxl.load_workbook("data.xlsx")
    st1=wb["Sheet1"]
    rows=st1.max_row
    for i in range(1,rows+1):
        if st1.cell(i,1).value == idval:
            break
    deleterow(i)
    for j in range(i,rows+1):
        st1.cell(j,1).value=st1.cell(j+1,1).value
        st1.cell(j,2).value=st1.cell(j+1,2).value
        st1.cell(j,3).value=st1.cell(j+1,3).value
        st1.cell(j,4).value=st1.cell(j+1,4).value
    wb.save("data.xlsx")
removebyid(4)
#deleterow(3)
#deleterow(4)
#new_admision("demo", "demo","demo")