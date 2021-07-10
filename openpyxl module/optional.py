import openpyxl
wb = openpyxl.load_workbook("data.xlsx")
st1=wb["Sheet1"]
print(type(st1))
print(st1.cell(1,1).value)
print(st1.max_row)
st1.cell(1,3).value="Mother Name"
wb.save("data.xlsx")
